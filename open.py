from openpyxl import load_workbook

# 엑셀 파일을 불러옵니다.
wb = load_workbook(filename='stock_info_test.xlsx')

# 모든 시트 이름을 가져옵니다.
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]  # 시트 선택
    
    # B11부터 K11까지의 값을 저장할 리스트를 초기화합니다.
    values_self_per = []

    # B11부터 K11까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(2, 12):  # B열은 2, K열은 11에 해당합니다.
        cell_value = sheet.cell(row=11, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_self_per.append(float(cell_value))
            except ValueError:
                continue

    # M14부터 Q14까지의 값을 저장할 리스트를 초기화합니다.
    values_section_per = []

    # M14부터 Q14까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(13, 18):  # M열은 13, Q열은 17에 해당합니다.
        cell_value = sheet.cell(row=14, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_section_per.append(float(cell_value))
            except ValueError:
                continue

    # B13부터 K13까지의 값을 저장할 리스트를 초기화합니다.
    values_self_pbr = []

    # B13부터 K13까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(2, 12):  # B열은 2, K열은 11에 해당합니다.
        cell_value = sheet.cell(row=13, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_self_pbr.append(float(cell_value))
            except ValueError:
                continue
            
    # M15부터 Q15까지의 값을 저장할 리스트를 초기화합니다.
    values_section_pbr = []

    # M15부터 Q15까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(13, 18):  # M열은 13, Q열은 17에 해당합니다.
        cell_value = sheet.cell(row=15, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_section_pbr.append(float(cell_value))
            except ValueError:
                continue

    # B6부터 K6까지의 값을 저장할 리스트를 초기화합니다.
    values_self_roe = []

    # B6부터 K6까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(2, 12):  # B열은 2, K열은 11에 해당합니다.
        cell_value = sheet.cell(row=6, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_self_roe.append(float(cell_value))
            except ValueError:
                continue   

    # M13부터 Q13까지의 값을 저장할 리스트를 초기화합니다.
    values_section_roe = []

    # M13부터 Q13까지의 셀 값을 순회하며 읽어옵니다.
    for col in range(13, 18):  # M열은 13, Q열은 17에 해당합니다.
        cell_value = sheet.cell(row=13, column=col).value
        if cell_value is not None:  # 셀 값이 존재하는 경우에만 처리
            try:
                values_section_roe.append(float(cell_value))
            except ValueError:
                continue        

    # Self PER 평균을 계산합니다.
    if values_self_per:
        average_self_per = sum(values_self_per) / len(values_self_per)
        print(f"Sheet '{sheet_name}' Self PER Average: {average_self_per}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Self PER average calculation.")

    # Section PER 평균을 계산합니다.
    if values_section_per:
        average_section_per = sum(values_section_per) / len(values_section_per)
        print(f"Sheet '{sheet_name}' Section PER Average: {average_section_per}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Section PER average calculation.")  

    # Self PBR 평균을 계산합니다.
    if values_self_pbr:
        average_self_pbr = sum(values_self_pbr) / len(values_self_pbr)
        print(f"Sheet '{sheet_name}' Self PER Average: {average_self_pbr}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Self PER average calculation.")

    # Section PBR 평균을 계산합니다.
    if values_section_pbr:
        average_section_pbr = sum(values_section_pbr) / len(values_section_pbr)
        print(f"Sheet '{sheet_name}' Section PBR Average: {average_section_pbr}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Section PBR average calculation.")  
    
    # Self ROE 평균을 계산합니다.
    if values_self_roe:
        average_self_roe = sum(values_self_roe) / len(values_self_roe)
        print(f"Sheet '{sheet_name}' Self ROE Average: {average_self_roe}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Self ROE average calculation.")

    # Section ROE 평균을 계산합니다.
    if values_section_roe:
        average_section_roe = sum(values_section_roe) / len(values_section_roe)
        print(f"Sheet '{sheet_name}' Section ROE Average: {average_section_roe}")
    else:
        print(f"Sheet '{sheet_name}' has no valid numeric data for Section ROE average calculation.")

    # 기본값으로 0 또는 적절한 값을 사용할 수 있습니다.
    average_self_per = average_self_per if average_self_per is not None else 0
    average_section_per = average_section_per if average_section_per is not None else 0
    average_section_pbr = average_section_pbr if average_section_pbr is not None else 0
    average_section_roe = average_section_roe if average_section_roe is not None else 0

    #시트에 데이터 추가
    sheet.append(['average_self_per', 'average_self_pbr', 'average_self_roe'])
    sheet.append([average_self_per, average_self_pbr, average_self_roe])
    sheet.append(['average_section_per', 'average_section_pbr', 'average_section_roe'])
    sheet.append([average_section_per, average_section_pbr, average_section_roe])

wb.save('stock_info.xlsx')